"""办公自动化服务

提供 Excel/PDF/CSV 文件处理、批量重命名、数据报告生成等
办公自动化能力。

特性:
    - Excel 读取与生成使用 openpyxl（延迟导入，不可用时降级）
    - PDF 文本提取使用 pdfplumber 或 PyPDF2（延迟导入，不可用时降级）
    - CSV 处理使用标准库 csv 模块
    - 批量重命名使用正则表达式
    - 数据报告生成支持自动汇总统计
"""

import csv
import io
import logging
import re
from datetime import datetime, timezone
from typing import Any

from app.core.exceptions import BusinessException, ValidationException

logger = logging.getLogger(__name__)


class OfficeService:
    """办公自动化服务类

    提供 Excel/PDF/CSV 文件转换、批量重命名、报告生成等
    办公自动化功能，所有方法均为异步方法。
    """

    @classmethod
    async def excel_to_json(cls, file_bytes: bytes) -> dict[str, Any]:
        """将 Excel 文件内容转为 JSON

        使用 openpyxl 读取 Excel 文件（延迟导入），
        解析第一个工作表，提取表头和数据行。

        参数:
            file_bytes: Excel 文件字节内容

        返回:
            包含 sheet_name、headers、rows、total_rows 的字典

        异常:
            ValidationException: 文件内容为空时
            BusinessException: Excel 解析失败或 openpyxl 未安装时
        """
        if not file_bytes:
            raise ValidationException("Excel 文件内容不能为空")

        try:
            # 延迟导入 openpyxl，避免强依赖
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl 未安装，无法处理 Excel 文件")
            raise BusinessException(message="Excel 处理依赖 openpyxl 未安装")

        try:
            # 从字节内容加载工作簿（data_only=True 读取计算后的值）
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
            sheet = wb.active
            sheet_name = sheet.title

            rows_iter = sheet.iter_rows(values_only=True)
            # 第一行作为表头
            try:
                header_row = next(rows_iter)
            except StopIteration:
                # 空表
                wb.close()
                return {
                    "sheet_name": sheet_name,
                    "headers": [],
                    "rows": [],
                    "total_rows": 0,
                }

            headers = [str(cell) if cell is not None else "" for cell in header_row]

            # 读取数据行
            data_rows: list[dict[str, Any]] = []
            for row in rows_iter:
                # 跳过全空行
                if all(cell is None for cell in row):
                    continue
                row_dict: dict[str, Any] = {}
                for idx, cell in enumerate(row):
                    key = headers[idx] if idx < len(headers) else f"column_{idx}"
                    row_dict[key] = cell
                data_rows.append(row_dict)

            wb.close()

            logger.info("Excel 转 JSON 完成: %s, 共 %d 行数据", sheet_name, len(data_rows))
            return {
                "sheet_name": sheet_name,
                "headers": headers,
                "rows": data_rows,
                "total_rows": len(data_rows),
            }

        except BusinessException:
            raise
        except Exception as e:
            logger.error("Excel 转 JSON 失败: %s", e, exc_info=True)
            raise BusinessException(message=f"Excel 解析失败: {e}")

    @classmethod
    async def json_to_excel(
        cls, data: list[dict], sheet_name: str = "Sheet1"
    ) -> bytes:
        """将 JSON 数据转为 Excel 文件字节

        使用 openpyxl 创建工作簿（延迟导入），
        将字典列表写入工作表。

        参数:
            data: JSON 数据列表，每个元素为一行数据
            sheet_name: 工作表名称，默认 Sheet1

        返回:
            Excel 文件的字节内容

        异常:
            BusinessException: Excel 生成失败或 openpyxl 未安装时
        """
        if not data:
            raise ValidationException("数据不能为空")

        try:
            # 延迟导入 openpyxl
            from openpyxl import Workbook
        except ImportError:
            logger.error("openpyxl 未安装，无法生成 Excel 文件")
            raise BusinessException(message="Excel 处理依赖 openpyxl 未安装")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name

            # 以第一个字典的键作为表头
            headers = list(data[0].keys())
            ws.append(headers)

            # 写入数据行
            for item in data:
                row = [item.get(h, "") for h in headers]
                ws.append(row)

            # 写入字节缓冲区
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            file_bytes = buffer.getvalue()
            wb.close()

            logger.info("JSON 转 Excel 完成: %s, 共 %d 行数据", sheet_name, len(data))
            return file_bytes

        except BusinessException:
            raise
        except Exception as e:
            logger.error("JSON 转 Excel 失败: %s", e, exc_info=True)
            raise BusinessException(message=f"Excel 生成失败: {e}")

    @classmethod
    async def merge_excel_files(cls, files: list[bytes]) -> bytes:
        """合并多个 Excel 文件到一个

        使用 openpyxl 读取多个 Excel 文件（延迟导入），
        将所有数据行合并到一个工作表中。
        以第一个文件的表头为合并后的表头。

        参数:
            files: Excel 文件字节内容列表

        返回:
            合并后的 Excel 文件字节内容

        异常:
            ValidationException: 文件列表为空时
            BusinessException: 合并失败或 openpyxl 未安装时
        """
        if not files:
            raise ValidationException("待合并文件列表不能为空")

        try:
            # 延迟导入 openpyxl
            from openpyxl import Workbook, load_workbook
        except ImportError:
            logger.error("openpyxl 未安装，无法合并 Excel 文件")
            raise BusinessException(message="Excel 处理依赖 openpyxl 未安装")

        try:
            merged_wb = Workbook()
            merged_ws = merged_wb.active
            merged_ws.title = "MergedSheet"

            headers_written = False
            total_rows = 0

            for file_bytes in files:
                if not file_bytes:
                    continue

                wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
                sheet = wb.active
                rows_iter = sheet.iter_rows(values_only=True)

                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    wb.close()
                    continue

                # 仅第一个文件的表头写入合并表
                if not headers_written:
                    merged_ws.append(list(header_row))
                    headers_written = True

                # 写入数据行
                for row in rows_iter:
                    if all(cell is None for cell in row):
                        continue
                    merged_ws.append(list(row))
                    total_rows += 1

                wb.close()

            buffer = io.BytesIO()
            merged_wb.save(buffer)
            buffer.seek(0)
            file_bytes = buffer.getvalue()
            merged_wb.close()

            logger.info("合并 %d 个 Excel 文件完成，共 %d 行数据", len(files), total_rows)
            return file_bytes

        except BusinessException:
            raise
        except Exception as e:
            logger.error("合并 Excel 文件失败: %s", e, exc_info=True)
            raise BusinessException(message=f"合并 Excel 文件失败: {e}")

    @classmethod
    async def pdf_extract_text(cls, file_bytes: bytes) -> dict[str, Any]:
        """提取 PDF 文本

        优先使用 pdfplumber 提取文本（延迟导入），
        不可用时降级到 PyPDF2，两者都不可用时返回空文本。

        参数:
            file_bytes: PDF 文件字节内容

        返回:
            包含 pages、text、page_texts 的字典

        异常:
            ValidationException: 文件内容为空时
            BusinessException: PDF 解析失败时
        """
        if not file_bytes:
            raise ValidationException("PDF 文件内容不能为空")

        # 优先尝试 pdfplumber
        try:
            from pdfplumber import open as pdfplumber_open
        except ImportError:
            pdfplumber_open = None
            logger.info("pdfplumber 未安装，尝试使用 PyPDF2")

        # 降级到 PyPDF2
        if pdfplumber_open is None:
            try:
                from PyPDF2 import PdfReader
            except ImportError:
                logger.error("pdfplumber 和 PyPDF2 均未安装，无法提取 PDF 文本")
                raise BusinessException(
                    message="PDF 文本提取需要 pdfplumber 或 PyPDF2，请安装其中之一"
                )

            try:
                reader = PdfReader(io.BytesIO(file_bytes))
                page_texts: list[str] = []
                for page in reader.pages:
                    text = page.extract_text() or ""
                    page_texts.append(text)
                full_text = "\n".join(page_texts)
                result = {
                    "pages": len(page_texts),
                    "text": full_text,
                    "page_texts": page_texts,
                }
                logger.info("PDF 文本提取完成（PyPDF2）: %d 页", len(page_texts))
                return result

            except BusinessException:
                raise
            except Exception as e:
                logger.error("PDF 文本提取失败（PyPDF2）: %s", e, exc_info=True)
                raise BusinessException(message=f"PDF 文本提取失败: {e}")

        # 使用 pdfplumber 提取
        try:
            with pdfplumber_open(io.BytesIO(file_bytes)) as pdf:
                page_texts = []
                for page in pdf.pages:
                    text = page.extract_text() or ""
                    page_texts.append(text)
                full_text = "\n".join(page_texts)
                result = {
                    "pages": len(page_texts),
                    "text": full_text,
                    "page_texts": page_texts,
                }
            logger.info("PDF 文本提取完成（pdfplumber）: %d 页", len(page_texts))
            return result

        except BusinessException:
            raise
        except Exception as e:
            logger.error("PDF 文本提取失败（pdfplumber）: %s", e, exc_info=True)
            raise BusinessException(message=f"PDF 文本提取失败: {e}")

    @classmethod
    async def batch_rename(
        cls, filenames: list[str], pattern: str, replacement: str
    ) -> dict[str, Any]:
        """批量重命名预览

        根据正则表达式模式对文件名进行替换，返回新旧文件名映射。
        仅生成预览，不实际修改文件。

        参数:
            filenames: 原始文件名列表
            pattern: 待替换的正则表达式模式
            replacement: 替换后的字符串

        返回:
            包含 mappings（新旧名称映射列表）和 total 的字典

        异常:
            ValidationException: 文件名列表为空或模式为空时
            BusinessException: 正则表达式无效时
        """
        if not filenames:
            raise ValidationException("文件名列表不能为空")
        if not pattern:
            raise ValidationException("替换模式不能为空")

        try:
            compiled = re.compile(pattern)
        except re.error as e:
            raise BusinessException(message=f"无效的正则表达式: {e}")

        try:
            mappings: list[dict[str, str]] = []
            for old_name in filenames:
                new_name = compiled.sub(replacement, old_name)
                mappings.append({"old_name": old_name, "new_name": new_name})

            logger.info("批量重命名预览完成: %d 个文件", len(mappings))
            return {"mappings": mappings, "total": len(mappings)}

        except BusinessException:
            raise
        except Exception as e:
            logger.error("批量重命名失败: %s", e, exc_info=True)
            raise BusinessException(message=f"批量重命名失败: {e}")

    @classmethod
    async def generate_report(cls, template_data: dict[str, Any]) -> dict[str, Any]:
        """生成数据报告

        根据模板数据生成汇总统计报告，包含记录总数、字段统计、
        数值型字段的统计信息（总和、均值、最大值、最小值）等。

        参数:
            template_data: 模板数据，需包含 title 和 data 字段
                           - title: 报告标题
                           - data: 数据列表（每个元素为字典）

        返回:
            包含 title、summary、generated_at 的字典

        异常:
            ValidationException: 标题或数据为空时
            BusinessException: 报告生成失败时
        """
        title = template_data.get("title", "")
        data = template_data.get("data", [])

        if not title:
            raise ValidationException("报告标题不能为空")
        if not data or not isinstance(data, list):
            raise ValidationException("报告数据不能为空且必须为列表")

        try:
            total_records = len(data)

            # 收集所有字段名
            all_fields: set[str] = set()
            for item in data:
                if isinstance(item, dict):
                    all_fields.update(item.keys())

            field_stats: dict[str, Any] = {}

            for field in all_fields:
                values = [
                    item[field]
                    for item in data
                    if isinstance(item, dict) and field in item and item[field] is not None
                ]

                if not values:
                    continue

                # 数值型字段统计
                numeric_values = [v for v in values if isinstance(v, (int, float))]
                if numeric_values:
                    field_stats[field] = {
                        "count": len(numeric_values),
                        "sum": sum(numeric_values),
                        "avg": round(sum(numeric_values) / len(numeric_values), 2),
                        "max": max(numeric_values),
                        "min": min(numeric_values),
                    }
                else:
                    # 非数值字段统计
                    field_stats[field] = {
                        "count": len(values),
                        "unique_count": len(set(str(v) for v in values)),
                    }

            summary: dict[str, Any] = {
                "total_records": total_records,
                "total_fields": len(all_fields),
                "field_names": sorted(all_fields),
                "field_stats": field_stats,
            }

            generated_at = datetime.now(timezone.utc).isoformat()

            logger.info("数据报告生成完成: %s, 共 %d 条记录", title, total_records)
            return {
                "title": title,
                "summary": summary,
                "generated_at": generated_at,
            }

        except BusinessException:
            raise
        except Exception as e:
            logger.error("生成报告失败: %s", e, exc_info=True)
            raise BusinessException(message=f"生成报告失败: {e}")

    @classmethod
    async def csv_to_json(cls, file_bytes: bytes) -> list[dict[str, Any]]:
        """CSV 转 JSON

        使用标准库 csv 模块解析 CSV 文本，
        第一行作为表头，后续行转为字典列表。

        参数:
            file_bytes: CSV 文件字节内容

        返回:
            转换后的字典列表

        异常:
            ValidationException: 文件内容为空时
            BusinessException: CSV 解析失败时
        """
        if not file_bytes:
            raise ValidationException("CSV 文件内容不能为空")

        try:
            # 尝试 UTF-8 解码，失败则尝试 GBK
            try:
                text = file_bytes.decode("utf-8")
            except UnicodeDecodeError:
                text = file_bytes.decode("gbk", errors="replace")

            reader = csv.DictReader(io.StringIO(text))
            data = [dict(row) for row in reader]

            logger.info("CSV 转 JSON 完成: %d 行数据", len(data))
            return data

        except BusinessException:
            raise
        except Exception as e:
            logger.error("CSV 转 JSON 失败: %s", e, exc_info=True)
            raise BusinessException(message=f"CSV 解析失败: {e}")

    @classmethod
    async def json_to_csv(cls, data: list[dict]) -> bytes:
        """JSON 转 CSV

        使用标准库 csv 模块将字典列表写入 CSV 格式，
        以第一个字典的键作为表头。

        参数:
            data: JSON 数据列表，每个元素为一行数据

        返回:
            CSV 文件的字节内容

        异常:
            ValidationException: 数据为空时
            BusinessException: CSV 生成失败时
        """
        if not data:
            raise ValidationException("数据不能为空")

        try:
            # 以第一个字典的键作为表头
            headers = list(data[0].keys())

            buffer = io.StringIO()
            writer = csv.DictWriter(buffer, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            for item in data:
                writer.writerow(item)

            csv_bytes = buffer.getvalue().encode("utf-8")
            buffer.close()

            logger.info("JSON 转 CSV 完成: %d 行数据", len(data))
            return csv_bytes

        except BusinessException:
            raise
        except Exception as e:
            logger.error("JSON 转 CSV 失败: %s", e, exc_info=True)
            raise BusinessException(message=f"CSV 生成失败: {e}")
